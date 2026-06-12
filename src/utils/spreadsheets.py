import csv
import io
import warnings
import zipfile
from xml.etree import ElementTree


def format_cell_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _cell_column_index(cell_ref):
    letters = ""
    for char in cell_ref:
        if char.isalpha():
            letters += char
        else:
            break
    if not letters:
        return 0

    index = 0
    for char in letters:
        index = index * 26 + (ord(char.upper()) - ord("A") + 1)
    return index - 1


def _shared_strings(archive):
    try:
        xml = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    root = ElementTree.fromstring(xml)
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings = []
    for item in root.findall("x:si", namespace):
        strings.append("".join(text.text or "" for text in item.findall(".//x:t", namespace)))
    return strings


def _workbook_sheets(archive):
    namespace = {
        "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    targets = {}
    for relationship in relationships.findall("rel:Relationship", namespace):
        targets[relationship.attrib["Id"]] = relationship.attrib["Target"]

    sheets = []
    for sheet in workbook.findall(".//x:sheet", namespace):
        relationship_id = sheet.attrib[f"{{{namespace['r']}}}id"]
        target = targets[relationship_id]
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        sheets.append((sheet.attrib["name"], target))
    return sheets


def _raw_xlsx_cell_value(cell, shared_strings, namespace):
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//x:t", namespace))

    value = cell.find("x:v", namespace)
    if value is None or value.text is None:
        return ""

    if cell_type == "s":
        index = int(value.text)
        return shared_strings[index] if index < len(shared_strings) else ""

    return value.text


def _raw_xlsx_rows(file_path):
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rows_by_sheet = []
    with zipfile.ZipFile(file_path) as archive:
        shared_strings = _shared_strings(archive)
        for sheet_name, sheet_path in _workbook_sheets(archive):
            root = ElementTree.fromstring(archive.read(sheet_path))
            rows = []
            for row in root.findall(".//x:sheetData/x:row", namespace):
                values_by_column = {}
                for cell in row.findall("x:c", namespace):
                    values_by_column[_cell_column_index(cell.attrib.get("r", ""))] = _raw_xlsx_cell_value(
                        cell,
                        shared_strings,
                        namespace,
                    )
                if values_by_column:
                    max_column = max(values_by_column)
                    rows.append([values_by_column.get(index, "") for index in range(max_column + 1)])
                else:
                    rows.append([])
            rows_by_sheet.append((sheet_name, rows))
    return rows_by_sheet


def _openpyxl_rows(file_path):
    import openpyxl

    rows_by_sheet = []
    wb = None
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Workbook contains no default style.*")
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_by_sheet.append((sheet_name, [list(row) for row in ws.iter_rows(values_only=True)]))
    finally:
        if wb is not None:
            wb.close()
    return rows_by_sheet


def _xls_rows(file_path):
    import xlrd

    rows_by_sheet = []
    wb = xlrd.open_workbook(str(file_path))
    for sheet in wb.sheets():
        rows = []
        for i in range(sheet.nrows):
            rows.append([sheet.cell_value(i, j) for j in range(sheet.ncols)])
        rows_by_sheet.append((sheet.name, rows))
    return rows_by_sheet


def _rows_have_values(rows_by_sheet):
    return any(
        any(any(format_cell_value(value) for value in row) for row in rows)
        for _, rows in rows_by_sheet
    )


def _spreadsheet_rows(file_path):
    if file_path.suffix.lower() == ".xls":
        return _xls_rows(file_path)

    rows_by_sheet = _openpyxl_rows(file_path)
    if not _rows_have_values(rows_by_sheet):
        rows_by_sheet = _raw_xlsx_rows(file_path)
    return rows_by_sheet


def rows_to_csv(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    for row_number, row in rows:
        writer.writerow([f"row_{row_number}", *[format_cell_value(value) for value in row]])
    return output.getvalue().strip()


def _find_header_context(rows, heading_row=None):
    if heading_row:
        for row_number, row in rows:
            if row_number == heading_row:
                return rows_to_csv([(row_number, row)])
        return ""

    for row_number, row in rows[:50]:
        joined = " ".join(format_cell_value(value).lower() for value in row)
        if "date" in joined and any(
            word in joined
            for word in ("amount", "debit", "credit", "withdraw", "deposit", "narration", "remark")
        ):
            return rows_to_csv([(row_number, row)])
    return ""


def excel_batches(file_path, rows_per_batch, heading_row=None, row_from=None, row_to=None):
    batches = []
    for sheet_name, sheet_rows in _spreadsheet_rows(file_path):
        numbered_rows = [
            (row_number, row)
            for row_number, row in enumerate(sheet_rows, start=1)
            if any(format_cell_value(value) for value in row)
        ]
        header_context = _find_header_context(numbered_rows, heading_row=heading_row)
        selected_row_from = row_from or ((heading_row + 1) if heading_row else None)
        selected_rows = [
            (row_number, row)
            for row_number, row in numbered_rows
            if (selected_row_from is None or row_number >= selected_row_from)
            and (row_to is None or row_number <= row_to)
        ]
        if not selected_rows:
            continue

        for start in range(0, len(selected_rows), rows_per_batch):
            chunk = selected_rows[start : start + rows_per_batch]
            batches.append(
                {
                    "label": f"{sheet_name} rows {chunk[0][0]}-{chunk[-1][0]}",
                    "content": rows_to_csv(chunk),
                    "header_context": header_context,
                }
            )
    return batches
